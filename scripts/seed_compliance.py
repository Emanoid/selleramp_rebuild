#!/usr/bin/env python3
"""
Seed Firestore from an Excel compliance tracker.

Re-running is safe: rows are upserted via deterministic document IDs derived
from key fields. New rows in the Excel are added; existing rows are updated
in place (status, notes, dates are preserved if already changed in the app).

Usage:
    python scripts/seed_compliance.py path/to/tracker.xlsx
    python scripts/seed_compliance.py path/to/tracker.xlsx --dry-run
    python scripts/seed_compliance.py path/to/tracker.xlsx --wipe
    python scripts/seed_compliance.py path/to/tracker.xlsx --service-account /path/to/sa.json
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Install openpyxl first:  pip install openpyxl")


# ── slug helpers ──────────────────────────────────────────────────────────────

def _slug(*parts: object) -> str:
    combined = "_".join(str(p).lower() for p in parts if p is not None)
    return re.sub(r"[^a-z0-9]+", "_", combined).strip("_")[:120]


def _parse_date(val: object) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s or s in ("-", "N/A", "ASAP"):
        return None
    for fmt in ("%b %d, %Y", "%B %d, %Y", "%Y-%m-%d", "%m/%d/%Y", "%b. %d, %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _normalise_status(raw: object) -> str:
    s = str(raw or "").strip()
    if "Done" in s or "✅" in s:
        return "Done"
    if "Overdue" in s or "⚠️" in s:
        return "Overdue"
    return "Pending"


# ── sheet readers ─────────────────────────────────────────────────────────────

def _iter_rows(ws, header_row: int):
    """Yield dicts keyed by header cell values, skipping blank rows."""
    headers = [ws.cell(row=header_row, column=c).value
               for c in range(1, ws.max_column + 1)]
    for r in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value
                  for c in range(1, ws.max_column + 1)]
        if not any(v is not None for v in values):
            continue
        yield dict(zip(headers, values))


def parse_quarterly(ws) -> list[dict]:
    records = []
    for row in _iter_rows(ws, header_row=4):
        filing_type = str(row.get("Filing Type") or "").strip()
        if not filing_type:
            continue
        year = row.get("Year")
        quarter = str(row.get("Quarter") or "").strip()
        assigned_raw = str(row.get("Assigned To") or "LLC").strip() or "LLC"
        assigned_to = [n.strip() for n in assigned_raw.split(",") if n.strip()] or ["LLC"]
        doc_id = _slug("q", year, filing_type, quarter)
        records.append({
            "_id": doc_id,
            "category": "quarterly",
            "filing_type": filing_type,
            "jurisdiction": str(row.get("Jurisdiction") or "").strip(),
            "year": int(year) if year else None,
            "period": f"{quarter} {year}".strip() if (quarter and year) else str(year or ""),
            "quarter": quarter or None,
            "due_date": _parse_date(row.get("Due Date")),
            "status": _normalise_status(row.get("Status")),
            "date_filed": _parse_date(row.get("Date Filed")),
            "confirmation_number": None,
            "assigned_to": assigned_to,
            "notes": str(row.get("Notes") or "").strip(),
        })
    return records


def parse_annual(ws) -> list[dict]:
    records = []
    for row in _iter_rows(ws, header_row=4):
        filing = str(row.get("Filing") or "").strip()
        if not filing:
            continue
        year = row.get("Year")
        assigned_raw = str(row.get("Assigned To") or "LLC").strip() or "LLC"
        assigned_to = [n.strip() for n in assigned_raw.split(",") if n.strip()] or ["LLC"]
        doc_id = _slug("a", year, filing)
        records.append({
            "_id": doc_id,
            "category": "annual",
            "filing_type": filing,
            "jurisdiction": str(row.get("Jurisdiction") or "").strip(),
            "year": int(year) if year else None,
            "period": str(year or ""),
            "quarter": None,
            "due_date": _parse_date(row.get("Due Date")),
            "status": _normalise_status(row.get("Status")),
            "date_filed": _parse_date(row.get("Date Filed")),
            "confirmation_number": None,
            "assigned_to": assigned_to,
            "notes": str(row.get("Notes") or "").strip(),
        })
    return records


def parse_one_time(ws) -> list[dict]:
    records = []
    for row in _iter_rows(ws, header_row=4):
        item = str(row.get("Item") or "").strip()
        if not item:
            continue
        assigned_raw = str(row.get("Assigned To") or "LLC").strip() or "LLC"
        assigned_to = [n.strip() for n in assigned_raw.split(",") if n.strip()] or ["LLC"]
        doc_id = _slug("ot", item)
        conf_raw = row.get("Confirmation #")
        records.append({
            "_id": doc_id,
            "category": "one_time",
            "filing_type": item,
            "jurisdiction": str(row.get("Jurisdiction") or "").strip(),
            "year": None,
            "period": "",
            "quarter": None,
            "due_date": _parse_date(row.get("Due / Target")),
            "status": _normalise_status(row.get("Status")),
            "date_filed": _parse_date(row.get("Date Completed")),
            "confirmation_number": str(conf_raw).strip() if conf_raw else None,
            "assigned_to": assigned_to,
            "notes": str(row.get("Notes") or "").strip(),
        })
    return records


# ── Firestore upsert ──────────────────────────────────────────────────────────

def _to_firestore_dt(d: date | None):
    if d is None:
        return None
    if isinstance(d, datetime):
        return d
    return datetime(d.year, d.month, d.day)


def upsert_all(db, records: list[dict]) -> tuple[int, int]:
    from firebase_admin import firestore as fs
    col = db.collection("filings")
    added = updated = 0
    for rec in records:
        doc_id = rec.pop("_id")
        rec["due_date"] = _to_firestore_dt(rec.get("due_date"))
        rec["date_filed"] = _to_firestore_dt(rec.get("date_filed"))
        rec["updated_at"] = fs.SERVER_TIMESTAMP

        ref = col.document(doc_id)
        exists = ref.get().exists
        if exists:
            structural = {k: v for k, v in rec.items()
                          if k not in ("status", "date_filed", "notes",
                                       "confirmation_number", "updated_by")}
            structural["updated_at"] = fs.SERVER_TIMESTAMP
            ref.update(structural)
            updated += 1
        else:
            rec.setdefault("created_at", fs.SERVER_TIMESTAMP)
            rec.setdefault("updated_by", "seed_script")
            ref.set(rec)
            added += 1
    return added, updated


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Seed Firestore from a compliance tracker Excel file"
    )
    parser.add_argument("xlsx", help="Path to the Excel tracker file")
    parser.add_argument(
        "--service-account",
        default="service_account.json",
        help="Path to Firebase service account JSON (default: service_account.json)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print what would be written without touching Firestore",
    )
    parser.add_argument(
        "--wipe",
        action="store_true",
        help="Delete ALL documents in the filings collection before seeding",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    print(f"Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    quarterly = parse_quarterly(wb["Quarterly"])
    annual = parse_annual(wb["Annual"])
    one_time = parse_one_time(wb["OneTime"])
    all_records = quarterly + annual + one_time

    print(
        f"Parsed  {len(quarterly)} quarterly + {len(annual)} annual + "
        f"{len(one_time)} one-time = {len(all_records)} total"
    )

    if args.dry_run:
        print("\n--- DRY RUN — no writes ---")
        for rec in all_records:
            print(
                f"  [{rec['category']:10s}] {rec['_id'][:60]:<60}  "
                f"{rec['status']}  due:{rec['due_date']}"
            )
        print(f"\nWould upsert {len(all_records)} documents. Remove --dry-run to apply.")
        return

    sa_path = Path(args.service_account)
    if not sa_path.exists():
        sys.exit(f"Service account not found: {sa_path}")

    try:
        import firebase_admin
        from firebase_admin import credentials, firestore
    except ImportError:
        sys.exit("Install firebase-admin:  pip install firebase-admin")

    if not firebase_admin._apps:
        cred = credentials.Certificate(str(sa_path))
        firebase_admin.initialize_app(cred)
    db = firestore.client()

    if args.wipe:
        print("Wiping filings collection…")
        col = db.collection("filings")
        deleted = 0
        while True:
            docs = list(col.limit(400).stream())
            if not docs:
                break
            for doc in docs:
                doc.reference.delete()
                deleted += 1
        print(f"Deleted {deleted} documents.")

    print("\nWriting to Firestore…")
    added, updated = upsert_all(db, all_records)
    print(f"Done.  Added: {added}   Updated (structural fields only): {updated}")


if __name__ == "__main__":
    main()
