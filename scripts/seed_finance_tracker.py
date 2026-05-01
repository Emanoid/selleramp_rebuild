#!/usr/bin/env python3
"""
Seed Firestore finance_tracker collections from CentralLineGroup_Finance_Tracker.xlsx.

Reads:
  Expenses sheet  → finance_expenses (rows 4+ where column A is "Recurring" or "Static")
  Amazon Income   → finance_income   (rows 4+ where column A is a datetime)

Re-running is safe: existing documents are overwritten in place.

Usage:
    python scripts/seed_finance_tracker.py                    # default xlsx in repo root
    python scripts/seed_finance_tracker.py path/to/file.xlsx
    python scripts/seed_finance_tracker.py --dry-run
    python scripts/seed_finance_tracker.py --service-account /path/to/sa.json
"""
from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Install openpyxl first:  pip install openpyxl")

# ── path setup ────────────────────────────────────────────────────────────────

_REPO = Path(__file__).resolve().parents[1]
_DEFAULT_XLSX = _REPO / "CentralLineGroup_Finance_Tracker.xlsx"

sys.path.insert(0, str(_REPO / "src"))


# ── helpers ───────────────────────────────────────────────────────────────────

def _to_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _to_iso(val) -> str | None:
    d = _to_date(val)
    return d.isoformat() if d else None


# ── Excel readers ─────────────────────────────────────────────────────────────

def _read_expenses(ws) -> list[dict]:
    """Parse Expenses sheet. Data rows start at row 4; col A = Type."""
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True)):
        exp_type = row[0]
        if exp_type not in ("Recurring", "Static"):
            continue
        freq = str(row[3]).strip() if row[3] else "Once"
        rows.append({
            "type":             exp_type,
            "item_description": str(row[1]).strip() if row[1] else "",
            "unit_price":       float(row[2]) if row[2] is not None else 0.0,
            "frequency":        freq,
            "start_date":       _to_iso(row[4]),
            "end_date":         _to_iso(row[5]),
            "receipt_filename": str(row[7]).strip() if row[7] else "",
            "notes":            str(row[8]).strip() if row[8] else "",
            "order":            i + 1,
        })
    return rows


def _read_income(ws) -> list[dict]:
    """Parse Amazon Income sheet. Data rows start at row 4; col A = Date."""
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True)):
        if not isinstance(row[0], (date, datetime)):
            continue
        rows.append({
            "date":   _to_iso(row[0]),
            "type":   str(row[1]).strip() if row[1] else "Income after Fees",
            "amount": float(row[2]) if row[2] is not None else 0.0,
            "notes":  str(row[3]).strip() if row[3] else "",
            "order":  i + 1,
        })
    return rows


# ── Firestore writer ──────────────────────────────────────────────────────────

def _write(db, collection: str, rows: list[dict], email: str, dry_run: bool) -> None:
    from google.cloud.firestore import SERVER_TIMESTAMP
    now = datetime.utcnow()

    for row in rows:
        doc_id = f"{collection}_{row['order']:04d}"
        payload = {
            **row,
            "created_at": now,
            "updated_at": now,
            "updated_by": email,
        }
        print(f"  {'[DRY RUN] ' if dry_run else ''}{collection}/{doc_id}: {row.get('item_description') or row.get('date')}")
        if not dry_run:
            db.collection(collection).document(doc_id).set(payload)


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", nargs="?", default=str(_DEFAULT_XLSX), help="Path to Excel file")
    ap.add_argument("--dry-run", action="store_true", help="Print rows without writing")
    ap.add_argument("--service-account", metavar="PATH", help="Firebase service-account JSON")
    ap.add_argument("--email", default="seed@script.local", help="updated_by email to record")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path)

    expenses = _read_expenses(wb["Expenses"])
    income   = _read_income(wb["Amazon Income"])

    print(f"\nFound {len(expenses)} expense row(s) and {len(income)} income row(s) in {xlsx_path.name}")

    if not args.dry_run:
        # Bootstrap Firebase
        import os
        os.environ.setdefault(
            "GOOGLE_APPLICATION_CREDENTIALS",
            args.service_account or str(_REPO / "firebase-service-account.json"),
        )
        try:
            from sa_rebuild.compliance.firebase_client import get_db
        except Exception as e:
            sys.exit(f"Firebase init failed: {e}")
        db = get_db()
    else:
        db = None

    print("\n--- Expenses ---")
    _write(db, "finance_expenses", expenses, args.email, args.dry_run)

    print("\n--- Income ---")
    _write(db, "finance_income", income, args.email, args.dry_run)

    print(f"\n{'[DRY RUN] ' if args.dry_run else ''}Done.")


if __name__ == "__main__":
    main()
