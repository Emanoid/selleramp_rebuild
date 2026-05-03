"""Streamlit page: Keepa SellerAmp Compare — offline FBA analysis from a Keepa export."""
from __future__ import annotations

import io
import sys
import tempfile
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_HERE = Path(__file__).resolve()
_SRC  = _HERE.parents[3]
if str(_SRC) not in sys.path:
    sys.path.insert(0, str(_SRC))

from sa_rebuild.keepa_compare.__main__ import run_compare  # noqa: E402

_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill("solid", fgColor="1565C0")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
_DATA_ALIGN   = Alignment(horizontal="center", vertical="center")


def _col_width(values: list, header: str) -> float:
    sample = [str(v) for v in values[:500]]
    return min(max(len(header), max((len(v) for v in sample), default=0)) + 2, 62)


def _to_xlsx(df: pd.DataFrame) -> bytes:
    """Styled Excel: blue headers, auto-filter, first column frozen."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Results")
        ws = writer.sheets["Results"]

        cols = list(df.columns)

        # Style header row
        for ci, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=ci)
            cell.font      = _HEADER_FONT
            cell.fill      = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN

        # Auto-filter + freeze first column
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
        ws.freeze_panes    = "B1"

        # Auto-fit column widths
        for ci, col in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = _col_width(
                df[col].tolist(), col
            )

    buf.seek(0)
    return buf.getvalue()


def _to_xlsx_transposed(df: pd.DataFrame) -> bytes:
    """Transposed styled Excel: fields as rows, products as columns.

    First column (field names) has the same blue styling as headers.
    Pane frozen at B2 so both the header row and field-name column stay fixed.
    """
    t_df = df.T.reset_index()
    t_df.columns = ["Field"] + [str(i) for i in range(len(df))]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        t_df.to_excel(writer, index=False, sheet_name="Transposed")
        ws = writer.sheets["Transposed"]

        ncols = len(t_df.columns)

        # Style header row (row 1)
        for ci in range(1, ncols + 1):
            cell = ws.cell(row=1, column=ci)
            cell.font      = _HEADER_FONT
            cell.fill      = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN

        # Style first data column (field names, column A, rows 2+)
        for ri in range(2, len(t_df) + 2):
            cell = ws.cell(row=ri, column=1)
            cell.font      = _HEADER_FONT
            cell.fill      = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN

        # Freeze both header row and field-name column
        ws.freeze_panes = "B2"

        # Auto-fit first column; fixed width for data columns
        ws.column_dimensions["A"].width = _col_width(t_df["Field"].tolist(), "Field")
        for ci in range(2, ncols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 14

    buf.seek(0)
    return buf.getvalue()


# ── Bold blue column headers (applied to every st.dataframe on this page) ────
st.markdown("""
<style>
[data-testid="stDataFrame"] thead tr th {
    color: #1565C0 !important;
    font-weight: 700 !important;
}
[data-testid="stDataFrame"] th[scope="row"] {
    color: #1565C0 !important;
    font-weight: 700 !important;
}
</style>
""", unsafe_allow_html=True)

st.title("📊 Keepa SellerAmp Compare")
st.markdown(
    "Match your wholesale cost list against a Keepa Product Viewer export — "
    "**no API key needed**. Get recommended sell price, ROI, profit, and fee breakdown "
    "for every product in seconds."
)

# ── Step 1: Template ──────────────────────────────────────────────────────────
st.markdown("---")
st.subheader("Step 1 — Download Input Template")
st.markdown(
    "Each row needs **either** a `upc` (barcode) **or** an `asin`, plus your "
    "wholesale `cost` and `weight_lbs`. Both `upc` and `asin` on the same row is fine too."
)
st.info(
    "**Excel tip:** format the `upc` column as **Plain Text** before pasting barcodes. "
    "Excel converts long numbers to scientific notation (`8.83503E+11`), which breaks matching."
)

_TEMPLATE = pd.DataFrame([
    {"upc": "701619100159", "asin": "",           "cost": 16.09, "weight_lbs": 0.28},
    {"upc": "840187704557", "asin": "",           "cost": 11.99, "weight_lbs": 0.10},
    {"upc": "",             "asin": "B08N5WRWNW", "cost": 8.50,  "weight_lbs": 0.55},
])
st.download_button(
    label="⬇ Download CSV Template",
    data=_TEMPLATE.to_csv(index=False),
    file_name="keepa_compare_template.csv",
    mime="text/csv",
)

# ── Step 2: How to export from Keepa ─────────────────────────────────────────
st.markdown("---")
st.subheader("Step 2 — Export from Keepa Product Viewer")

with st.expander("How to get the Keepa export file (click to expand)"):
    st.markdown("""
1. Go to [keepa.com](https://keepa.com) and open **Product Viewer** (top navigation bar).
2. In the search/import box, paste your UPCs, EANs, or ASINs — **comma separated** (e.g. `701619100159, 840187704557, B08N5WRWNW`).
3. Wait for the results table to fully load (all rows should show price data).
4. Click the **Export** button (top-right of the results grid).
5. Choose **Excel (.xlsx)** or **CSV** — both work here.
6. Save the file and upload it in Step 3 below.

> The key column Keepa exports is called **"Imported by Code"** — it records what code
> you searched for each product (UPC, EAN, ASIN). This is what links your wholesale list
> back to the Keepa data.
""")

# ── Step 3: Upload files ──────────────────────────────────────────────────────
st.markdown("---")
st.subheader("Step 3 — Upload Your Files")

col_left, col_right = st.columns(2)

with col_left:
    st.markdown("**Your Cost / Product List**")
    st.caption("CSV with columns: `upc`, `asin`, `cost`, `weight_lbs`")
    cost_file = st.file_uploader(
        "Upload cost CSV", type=["csv"],
        key="cost_upload", label_visibility="collapsed",
    )

with col_right:
    st.markdown("**Keepa Product Viewer Export**")
    st.caption('Excel (.xlsx) or CSV exported from Keepa. Must include "Imported by Code" column.')
    keepa_file = st.file_uploader(
        "Upload Keepa export", type=["xlsx", "xls", "csv"],
        key="keepa_upload", label_visibility="collapsed",
    )

# ── Step 4: Run ───────────────────────────────────────────────────────────────
st.markdown("---")
st.subheader("Step 4 — Run Comparison")

if cost_file and keepa_file:
    if st.button("▶ Run Comparison", type="primary"):
        with st.spinner("Matching products and computing FBA metrics…"):
            with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as tf_c:
                tf_c.write(cost_file.read())
                cost_tmp = Path(tf_c.name)

            keepa_suffix = Path(keepa_file.name).suffix
            with tempfile.NamedTemporaryFile(suffix=keepa_suffix, delete=False) as tf_k:
                tf_k.write(keepa_file.read())
                keepa_tmp = Path(tf_k.name)

            try:
                result_df = run_compare(cost_tmp, keepa_tmp)
            except Exception as exc:
                st.error(f"Error during comparison: {exc}")
                cost_tmp.unlink(missing_ok=True)
                keepa_tmp.unlink(missing_ok=True)
                st.stop()
            finally:
                cost_tmp.unlink(missing_ok=True)
                keepa_tmp.unlink(missing_ok=True)

        with st.spinner("Building Excel files…"):
            xlsx_bytes   = _to_xlsx(result_df)
            xlsx_t_bytes = _to_xlsx_transposed(result_df)
            rec_df       = result_df[result_df.get("Investment Decision") == "Recommend"]
            xlsx_rec_bytes = _to_xlsx_transposed(rec_df)

        st.session_state["kc_result"]    = result_df
        st.session_state["kc_xlsx"]      = xlsx_bytes
        st.session_state["kc_xlsx_t"]    = xlsx_t_bytes
        st.session_state["kc_xlsx_rec"]  = xlsx_rec_bytes
        st.session_state["kc_summary"]  = {
            "total":     len(result_df),
            "recommend": int((result_df.get("Investment Decision", pd.Series(dtype=str)) == "Recommend").sum()),
            "no_match":  int((result_df.get("Title", pd.Series(dtype=str)) == "NO MATCH FOUND").sum()),
            "not_rec":   int((result_df.get("Investment Decision", pd.Series(dtype=str)) == "Not Recommend").sum()),
        }
else:
    st.info("Upload both files above to enable the comparison.")

# ── Display results ───────────────────────────────────────────────────────────
if "kc_result" in st.session_state:
    summary    = st.session_state["kc_summary"]
    result_df  = st.session_state["kc_result"]

    not_rec_shown = summary["not_rec"] - summary["no_match"]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Rows",       summary["total"])
    c2.metric("✅ Recommend",      summary["recommend"])
    c3.metric("❌ Not Recommend",  not_rec_shown)
    c4.metric("⚠ No Match",        summary["no_match"])

    # ── Download buttons — above the preview so they appear immediately ───────
    dl1, dl2, dl3, dl4 = st.columns(4)
    with dl1:
        st.download_button(
            label="⬇ Download Excel",
            data=st.session_state["kc_xlsx"],
            file_name="keepa_compare_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with dl2:
        st.download_button(
            label="⬇ Download Excel (Transposed)",
            data=st.session_state["kc_xlsx_t"],
            file_name="keepa_compare_transposed.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with dl3:
        st.download_button(
            label="⬇ Download CSV",
            data=result_df.to_csv(index=False),
            file_name="keepa_compare_output.csv",
            mime="text/csv",
        )
    with dl4:
        st.download_button(
            label="✅ Download Recommended Only",
            data=st.session_state["kc_xlsx_rec"],
            file_name="keepa_compare_recommended.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.dataframe(result_df, use_container_width=True, height=500)
